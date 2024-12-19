from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import logging

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


class Exporter:
    """
    A class to export data into an Excel template and create a new file.
    Highlights rows in red if a specific column (e.g., AG) is empty and other values are present.
    """

    def __init__(self, config: dict):
        """
        Initializes the Exporter with a configuration.

        :param config: Configuration dictionary.
        """
        self.template_path = config["config_templates"]["path"]["bawa"]
        self.cell_mapping = config["cell_mapping"]
        self.highlight_color = config.get("highlight", {}).get("empty_ag_color", "FF0000")  # Default red
        logger.info("Exporter initialized with template path and output configuration.")

    def _get_column_and_start_row(self, sheet_mapping: dict, target_key: str) -> tuple:
        """
        Finds the column and starting row for a specific key in the sheet mapping.

        :param sheet_mapping: Dictionary of cell-to-key mapping for a sheet.
        :param target_key: The key to search for (e.g., "cvlan").
        :return: Tuple (column, start_row).
        """
        for cell, key in sheet_mapping.items():
            if key == target_key:
                column = ''.join(filter(str.isalpha, cell))  # Extract letters (e.g., "AG")
                start_row = int(''.join(filter(str.isdigit, cell)))  # Extract digits (e.g., 14)
                return column, start_row
        raise ValueError(f"Column for '{target_key}' not found in the cell mapping.")

    @staticmethod
    def _highlight_row(sheet, row: int, color: str):
        """
        Highlights an entire row with a specified color.

        :param sheet: The worksheet object.
        :param row: Row number to highlight.
        :param color: Color code for highlighting (hex format, e.g., "FF0000").
        """
        red_fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            cell.fill = red_fill

    def export(self, data: dict, output_path: str) -> None:
        """
        Exports data into the Excel template.

        :param data: Dictionary containing the data to export.
        :param output_path: Path to save the resulting Excel file.
        """
        logger.info("Loading Excel template...")
        workbook = load_workbook(self.template_path)

        for sheet_name, sheet_mapping in self.cell_mapping.items():
            if sheet_name not in workbook.sheetnames:
                logger.error(f"Sheet '{sheet_name}' not found in the template.")
                continue

            sheet = workbook[sheet_name]
            logger.info(f"Processing sheet: {sheet_name}")

            try:
                ag_column, start_row = self._get_column_and_start_row(sheet_mapping, "cvlan")
            except ValueError as e:
                logger.error(e)
                continue

            current_row = start_row

            for row_data in data.values():
                # Checking if the line is filled
                while any(sheet[f"{col}{current_row}"].value is not None
                          for col in [cell[:1] for cell in sheet_mapping.keys()]):
                    current_row += 1

                # Inserting data into cells
                row_filled = False
                for cell, key in sheet_mapping.items():
                    column_letter = ''.join(filter(str.isalpha, cell))
                    if key in row_data and row_data[key] is not None:
                        sheet[f"{column_letter}{current_row}"] = row_data[key]
                        row_filled = True

                # String highlighting if cvlan (AG) is empty, but other cells are populated
                ag_cell = sheet[f"{ag_column}{current_row}"]
                if ag_cell.value is None and row_filled:
                    logger.warning(f"Highlighting row {current_row} due to empty '{ag_column}' column.")
                    self._highlight_row(sheet, current_row, self.highlight_color)

                current_row += 1

        # Сохранение Excel-файла
        workbook.save(output_path)
        logger.info(f"Excel file successfully saved at {output_path}.")
